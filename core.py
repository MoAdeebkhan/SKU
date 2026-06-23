"""
core.py
-------
Shared logic for stock balancing. No AI, no UI — pure data processing.
Used by both the with-AI and without-AI versions.
"""

import io
from collections import defaultdict
from dataclasses import dataclass, field

import openpyxl
import pandas as pd


# ---------------------------------------------------------------------------
# Column name constants — change here if your Excel headers differ
# ---------------------------------------------------------------------------
COL_MATERIAL = "Material Number"
COL_SKU      = "SKU Name"
COL_LOCATION = "Receiving Location"
COL_STOCK    = "Number of Stocks"

REQUIRED_COLUMNS = [COL_MATERIAL, COL_SKU, COL_LOCATION, COL_STOCK]


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class SkuGroup:
    """Holds data for one SKU that appears in multiple locations."""
    sku_name:   str
    df_indices: list       # original DataFrame row indices
    locations:  list[str]
    old_stocks: list[int]
    new_stocks: list[int]  = field(default_factory=list)
    total:      int        = 0

    def is_already_balanced(self) -> bool:
        return len(set(self.old_stocks)) == 1

    def variance(self) -> int:
        return max(self.old_stocks) - min(self.old_stocks)


@dataclass
class BalancingResult:
    """Full result of one balancing run."""
    sku_groups:      list[SkuGroup]
    total_rows:      int
    total_skus:      int
    unique_skus:     int
    skus_balanced:   int
    rows_updated:    int
    total_stock:     int


# ---------------------------------------------------------------------------
# Excel loading
# ---------------------------------------------------------------------------

class ExcelLoadError(Exception):
    pass


def load_excel_bytes(file_bytes: bytes) -> tuple[pd.DataFrame, openpyxl.Workbook]:
    """
    Load an Excel file from raw bytes.

    Returns (DataFrame, Workbook).
    Raises ExcelLoadError on any problem.

    Edge cases handled:
      - Empty file
      - Missing required columns
      - Non-numeric stock values (coerced to 0)
      - Completely blank rows
      - Whitespace-only SKU names
      - Duplicate column names in header
    """
    if not file_bytes:
        raise ExcelLoadError("The uploaded file is empty.")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), keep_vba=True)
    except Exception as exc:
        raise ExcelLoadError(f"Cannot open file: {exc}")

    ws = wb.active
    if ws is None:
        raise ExcelLoadError("The workbook has no active sheet.")

    # Read header row
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None or all(v is None for v in header_row):
        raise ExcelLoadError("The first row (header) is empty.")

    # Build header map; detect duplicate column names
    seen = {}
    headers = {}
    for col_idx, val in enumerate(header_row, start=1):
        if val is None:
            continue
        name = str(val).strip()
        if name in seen:
            raise ExcelLoadError(
                f"Duplicate column name '{name}' found at columns "
                f"{seen[name]} and {col_idx}."
            )
        seen[name] = col_idx
        headers[name] = col_idx

    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        found = list(headers.keys())
        raise ExcelLoadError(
            f"Required column(s) not found: {missing}\n"
            f"Columns detected in file: {found}"
        )

    # Read all data rows into a list of dicts
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for col_name, col_idx in headers.items():
            row_dict[col_name] = row[col_idx - 1]
        data.append(row_dict)

    if not data:
        raise ExcelLoadError("The file has a header but no data rows.")

    df = pd.DataFrame(data)

    # Drop rows where every cell is None
    df.dropna(how="all", inplace=True)

    # Normalise SKU column: strip whitespace, drop blank/None
    df[COL_SKU] = df[COL_SKU].astype(str).str.strip()
    df = df[df[COL_SKU].notna() & (df[COL_SKU] != "") & (df[COL_SKU] != "None")]

    if df.empty:
        raise ExcelLoadError("No valid data rows found after removing blank rows.")

    # Coerce stock to int; invalid values become 0
    df[COL_STOCK] = pd.to_numeric(df[COL_STOCK], errors="coerce").fillna(0).astype(int)

    # Negative stock is treated as 0
    df[COL_STOCK] = df[COL_STOCK].clip(lower=0)

    # Reset index so it matches original file row positions
    df.reset_index(drop=True, inplace=True)

    return df, wb


# ---------------------------------------------------------------------------
# Duplicate detection
# ---------------------------------------------------------------------------

def find_duplicate_skus(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """
    Return {sku_name: sub-DataFrame} for every SKU that appears
    in more than one row.

    Edge cases:
      - Case-sensitive matching (SKU-A and sku-a are different)
      - SKU with 1 location → excluded
      - All stocks zero → still returned so they get balanced to zero
    """
    result = {}
    for sku_name, grp in df.groupby(COL_SKU, sort=False):
        if len(grp) > 1:
            result[str(sku_name)] = grp
    return result


# ---------------------------------------------------------------------------
# Distribution algorithm
# ---------------------------------------------------------------------------

def compute_distribution(total: int, n: int) -> list[int]:
    """
    Distribute `total` across `n` slots with no decimals.

    Rules:
      - base      = total // n
      - remainder = total % n
      - `remainder` slots get (base + 1)
      - remaining slots get base
      - sum(result) == total always

    Edge cases:
      - total == 0  → all zeros
      - n == 1      → [total]
      - total < n   → some slots get 1, rest get 0
    """
    if n <= 0:
        raise ValueError(f"Number of locations must be >= 1, got {n}")
    if total < 0:
        raise ValueError(f"Total stock cannot be negative, got {total}")
    if n == 1:
        return [total]
    if total == 0:
        return [0] * n

    base      = total // n
    remainder = total % n
    dist = [base + 1] * remainder + [base] * (n - remainder)
    assert sum(dist) == total, f"Distribution bug: sum={sum(dist)}, expected={total}"
    return dist


# ---------------------------------------------------------------------------
# Build SkuGroup objects
# ---------------------------------------------------------------------------

def build_sku_groups(duplicates: dict[str, pd.DataFrame]) -> list[SkuGroup]:
    """
    Convert duplicate-SKU DataFrames into SkuGroup objects
    with new_stocks already computed.
    """
    groups = []
    for sku_name, grp in duplicates.items():
        old_stocks = grp[COL_STOCK].tolist()
        locations  = [str(v) for v in grp[COL_LOCATION].tolist()]
        total      = sum(old_stocks)
        new_stocks = compute_distribution(total, len(old_stocks))

        groups.append(SkuGroup(
            sku_name   = sku_name,
            df_indices = list(grp.index),
            locations  = locations,
            old_stocks = old_stocks,
            new_stocks = new_stocks,
            total      = total,
        ))
    return groups


# ---------------------------------------------------------------------------
# Apply balancing to workbook
# ---------------------------------------------------------------------------

def apply_balancing_to_workbook(
    wb: openpyxl.Workbook,
    sku_groups: list[SkuGroup],
) -> bytes:
    """
    Write new stock values into the workbook and return it as bytes.

    The DataFrame index (0-based) maps to Excel row index+2
    (row 1 = header, rows start at 2).

    Edge cases:
      - Preserves all other columns untouched
      - Preserves VBA macros (keep_vba=True was set on load)
      - Returns bytes so the original file object is not mutated
    """
    ws = wb.active

    # Rebuild header map from the live workbook
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {
        str(val).strip(): col_idx
        for col_idx, val in enumerate(header_row, start=1)
        if val is not None
    }
    stock_col = headers[COL_STOCK]

    for group in sku_groups:
        for df_idx, new_val in zip(group.df_indices, group.new_stocks):
            excel_row = df_idx + 2  # df index 0 -> Excel row 2
            ws.cell(row=excel_row, column=stock_col).value = new_val

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Rule-based analysis text (no-AI version)
# ---------------------------------------------------------------------------

def generate_analysis_text(sku_groups: list[SkuGroup]) -> str:
    """
    Generate a plain-English analysis report from the data alone.
    No model required.
    """
    if not sku_groups:
        return "No SKUs required rebalancing."

    lines = []
    lines.append(
        f"Analysis found {len(sku_groups)} SKU(s) distributed across "
        f"multiple warehouse locations with unequal stock levels."
    )
    lines.append("")

    for g in sku_groups:
        if g.is_already_balanced():
            lines.append(
                f"  {g.sku_name}: Already balanced at {g.old_stocks[0]} units "
                f"per location across {len(g.locations)} location(s). No change needed."
            )
        else:
            lines.append(
                f"  {g.sku_name}: Total stock {g.total} units spread unevenly across "
                f"{len(g.locations)} location(s) (range: {min(g.old_stocks)} to "
                f"{max(g.old_stocks)}). Rebalanced to approximately "
                f"{g.new_stocks[0]} units per location."
            )

    lines.append("")
    lines.append(
        "Rebalancing ensures equal stock availability across all warehouse "
        "locations, reducing the risk of stockouts at low-inventory sites "
        "and overstocking at high-inventory sites."
    )
    return "\n".join(lines)


def generate_summary_text(sku_groups: list[SkuGroup], dry_run: bool) -> str:
    """
    Generate a short operation summary from the data alone.
    """
    total_rows    = sum(len(g.locations) for g in sku_groups)
    total_units   = sum(g.total for g in sku_groups)
    mode_str      = "Preview (dry run)" if dry_run else "Applied"

    return (
        f"{mode_str}: Rebalanced {len(sku_groups)} SKU(s) across "
        f"{total_rows} location row(s), covering {total_units:,} total stock units. "
        f"Each SKU's stock is now distributed as evenly as possible, "
        f"with any indivisible remainder assigned one unit at a time to the first locations."
    )


# ---------------------------------------------------------------------------
# Build BalancingResult summary
# ---------------------------------------------------------------------------

def build_result(
    df: pd.DataFrame,
    sku_groups: list[SkuGroup],
) -> BalancingResult:
    return BalancingResult(
        sku_groups    = sku_groups,
        total_rows    = len(df),
        total_skus    = df[COL_SKU].nunique(),
        unique_skus   = df[COL_SKU].nunique(),
        skus_balanced = len(sku_groups),
        rows_updated  = sum(len(g.locations) for g in sku_groups),
        total_stock   = int(df[COL_STOCK].sum()),
    )